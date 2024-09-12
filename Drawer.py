from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import matplotlib.font_manager as fm
import platform

class Drawer:
    def __init__(self, output_directory, student_name, student_id, student_class, kebiao_df, room_df, schedule_format_directory, platform):
        self.output_directory = output_directory
        self.student_name = student_name
        self.student_id = student_id
        self.student_class = student_class
        self.kebiao_df = kebiao_df
        self.room_df = room_df
        self.schedule_format_directory = schedule_format_directory
        self.template_schedule = load_workbook(self.schedule_format_directory)
        self.template_schedule_sheet = self.template_schedule.active

        self.platform = platform

        # 模版区块数据
        self.template_indexs = self.getTemplateIndexs()

    def getTemplateIndexs(self):
        # 判断新老模版 简单方法未来这部分要优化
        if 'old' in self.schedule_format_directory:
            return {
                "MON": ["C", "E"],
                "TUE": ["F", "H"],
                "WED": ["I", "K"],
                "THU": ["L", "N"],
                "FRI": ["O", "Q"],
                0: [8, 19],
                1: [8, 19],
                2: [20, 31],
                3: [20, 31],
                4: [35, 46],
                5: [35, 46],
                6: [47, 52]
            }
        else:
            return {
                "MON": ["C", "E"],
                "TUE": ["F", "H"],
                "WED": ["I", "K"],
                "THU": ["L", "N"],
                "FRI": ["O", "Q"],
                0: [8, 19],
                1: [8, 19],
                2: [20, 31],
                3: [20, 31],
                4: [37, 48],
                5: [37, 48],
                6: [49, 54]
            }

    def unmergeCell(self, a, b):
        self.template_schedule_sheet.unmerge_cells(f'{a}:{b}')
        self.template_schedule_sheet.merge_cells(f'{a}:{b[0:1]}{(int(b[1:len(b)]) - 6)}')
        self.template_schedule_sheet.merge_cells(f'{a[0:1]}{(int(a[1:len(a)]) + 6)}:{b}')

    def stnameAndClass(self):
        self.template_schedule_sheet["A1"].value = str(self.student_name) + ' - ' + str(self.student_class)
        # self.template_schedule_sheet["A5"].value = self.student_class

    def set_cell_border(self, cell, border_style):
        side = Side(border_style=border_style, color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        cell.border = border

    def find_font(self, font_name, default_font="Arial"):
        # 获取系统中所有可用字体的完整路径
        available_fonts = fm.findSystemFonts(fontpaths=None, fontext='ttf')

        for font_path in available_fonts:
            try:
                font_prop = fm.FontProperties(fname=font_path)
                if font_name.lower() in font_prop.get_name().lower():
                    return font_prop.get_name()
            except RuntimeError as e:
                print(f"无法处理字体 {font_path}: {e}")
                continue  # 跳过该字体并继续处理下一个

        # 如果未找到指定字体，返回默认字体名称
        return default_font

# 设置外边框
    def Draw(self, format):

        self.stnameAndClass()

        font_name = self.find_font("微软雅黑")
        if font_name != '微软雅黑':
            font_name = self.find_font("Microsoft YaHei")
        print(f"使用的字体名称: {font_name}")

        # 创建样式
        my_style = NamedStyle(name="custom_style")
        my_style.font = Font(name=font_name, size=22, bold=False)
        my_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        side = Side(border_style='medium', color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        my_style.border = border

        for day in ["MON", "TUE", "WED", "THU", "FRI"]:
            index = 0
            while(index < 7):
                continues = False
                if(index <= 4 and index % 2 == 0):
                    continues = (self.kebiao_df[day][index] == self.kebiao_df[day][index + 1])

                if(not continues and (index <= 4 and index % 2 == 0)):
                    try:
                        self.unmergeCell(self.template_indexs[day][0] + str(self.template_indexs[index][0]),
                                         self.template_indexs[day][1] + str(self.template_indexs[index][1]))
                    except Exception as e:
                        print(e)

                    cell = self.template_schedule_sheet[
                        f"{self.template_indexs[day][0]}{str(self.template_indexs[index][0] + 6)}"]
                    cell.value = str(self.kebiao_df[day][index + 1]) + '\n' + '(' + str(self.room_df[day][index + 1]) + ')'
                    cell.style = my_style
                    # 设置边框
                    self.set_cell_border(cell, "medium")

                cell = self.template_schedule_sheet[
                    f"{self.template_indexs[day][0]}{str(self.template_indexs[index][0])}"]
                cell.value = str(self.kebiao_df[day][index]) + '\n' + '(' + str(self.room_df[day][index]) + ')'
                cell.style = my_style
                self.set_cell_border(cell, "medium")

                index += 2

        #self.apply_border_to_range(self.template_schedule, 'A1', 'S60')

        if(format == 'Excel'):
            self.template_schedule.save(
                f"{self.output_directory}/{self.student_id}_{self.student_name}_{self.student_class}_ASE_schedule.xlsx")
        elif(format == '图片'):
            if (self.platform == 'Windows'):
                raise Exception('我还没写出来')
            else:
                raise Exception('当前系统不支持PDF格式')
        elif(format == 'PDF'):
            if(self.platform == 'Windows'):
                raise Exception('我还没写出来')
            else:
                raise Exception('当前系统不支持PDF格式')
        else:
            raise Exception("代码内无支持格式")




